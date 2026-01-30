"""
Excel handler for reading and writing invoice data
"""

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import TEMPLATE_FILE, INVOICE_FIELDS, OUTPUT_FOLDER


class ExcelHandler:
    """Handle Excel operations for invoice template"""
    
    def __init__(self, template_path=TEMPLATE_FILE):
        """Initialize with template path"""
        self.template_path = template_path
        self.workbook = None
        self.worksheet = None
        
    def load_template(self):
        """Load the template Excel file"""
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
            self.workbook = load_workbook(self.template_path)
            # Get the first sheet or 'Invoice' sheet
            sheet_name = 'Invoice' if 'Invoice' in self.workbook.sheetnames else self.workbook.sheetnames[0]
            self.worksheet = self.workbook[sheet_name]
            return True
        except Exception as e:
            raise Exception(f"Error loading template: {str(e)}")
    
    def get_cell_value(self, cell_ref):
        """Get value from a specific cell"""
        try:
            return self.worksheet[cell_ref].value
        except Exception as e:
            raise Exception(f"Error reading cell {cell_ref}: {str(e)}")
    
    def set_cell_value(self, cell_ref, value):
        """Set value to a specific cell"""
        try:
            self.worksheet[cell_ref].value = value
        except Exception as e:
            raise Exception(f"Error writing to cell {cell_ref}: {str(e)}")
    
    def get_all_template_values(self):
        """Get all template field values"""
        values = {}
        try:
            for field_key, field_config in INVOICE_FIELDS.items():
                cell_ref = field_config['cell']
                try:
                    values[field_key] = self.get_cell_value(cell_ref)
                except:
                    values[field_key] = ''
            return values
        except Exception as e:
            raise Exception(f"Error reading template values: {str(e)}")
    
    def update_invoice(self, data_dict):
        """Update invoice with provided data"""
        try:
            for field_key, value in data_dict.items():
                if field_key in INVOICE_FIELDS:
                    field_config = INVOICE_FIELDS[field_key]
                    if not field_config.get('read_only', False):
                        cell_ref = field_config['cell']
                        self.set_cell_value(cell_ref, value)
        except Exception as e:
            raise Exception(f"Error updating invoice: {str(e)}")
    
    def save_invoice(self, output_filename=None):
        """Save the modified invoice"""
        try:
            # Create output folder if it doesn't exist
            if not os.path.exists(OUTPUT_FOLDER):
                os.makedirs(OUTPUT_FOLDER)
            
            if output_filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"Invoice_{timestamp}.xlsx"
            
            output_path = os.path.join(OUTPUT_FOLDER, output_filename)
            self.workbook.save(output_path)
            return output_path
        except Exception as e:
            raise Exception(f"Error saving invoice: {str(e)}")
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()
